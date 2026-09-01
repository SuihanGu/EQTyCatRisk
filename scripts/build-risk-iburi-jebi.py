"""
将第二页算例（2018 胆振东部地震 × 台风21号飞燕）转为前端 JSON。
用法: python scripts/build-risk-iburi-jebi.py
输出: public/data/risk-iburi-jebi.json

网格损失来自 Coupled_Loss_with_Other_JPY（单位 JPY）。
"""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "第二页数据"
OUT = ROOT / "public" / "data" / "risk-iburi-jebi.json"
OUT_GRID = ROOT / "public" / "data" / "risk-grid-cells.json"

EQ_XLSX = DATA / "地震信息.xlsx"
TRACK_XLSX = DATA / "21号飞燕的台风轨迹.xlsx"
EXPOSURE_CSV = DATA / "第二页_北海道风险暴露数据_PGA_Zhao2016方法_含风速（新）.csv"

# 网格边长约 30"（0.008333°）；半宽用于前端画矩形
GRID_HALF_DEG = 0.0041666667

STRUCTURE_SHARE = {"wood": 0.42, "steel": 0.18, "rc": 0.28, "masonry": 0.12}

WKT_NUM = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?")


def to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def open_csv(path: Path):
    """新表含日文列名，优先 gbk，再试 utf-8-sig。"""
    last_err: Exception | None = None
    for enc in ("gbk", "utf-8-sig", "cp932"):
        try:
            f = path.open("r", encoding=enc, newline="")
            reader = csv.DictReader(f)
            # 触发解码
            _ = reader.fieldnames
            return f, reader, enc
        except Exception as err:
            last_err = err
            try:
                f.close()  # type: ignore[name-defined]
            except Exception:
                pass
    raise RuntimeError(f"无法读取 {path}: {last_err}")


def load_earthquake() -> dict:
    rows = list(openpyxl.load_workbook(EQ_XLSX, data_only=True).active.iter_rows(values_only=True))
    _, lon, lat, depth, mag = rows[1]
    t = rows[1][0]
    if isinstance(t, datetime):
        eq_time = t.strftime("%Y-%m-%dT%H:%M:00+09:00")
        year = t.year
    else:
        eq_time = str(t)
        year = 2018
    return {
        "time": eq_time,
        "lon": float(lon),
        "lat": float(lat),
        "depthKm": float(depth),
        "magnitude": float(mag),
        "year": year,
    }


def load_typhoon_track() -> tuple[list[dict], float]:
    rows = list(openpyxl.load_workbook(TRACK_XLSX, data_only=True).active.iter_rows(values_only=True))
    path: list[dict] = []
    peak_kt = 0.0
    for row in rows[1:]:
        year, month, day, hour, _num, name, _grade, lat, lon, pres, vmax, *_rest = row
        lat_f, lon_f = to_float(lat), to_float(lon)
        if lat_f is None or lon_f is None:
            continue
        vmax_f = to_float(vmax) or 0.0
        peak_kt = max(peak_kt, vmax_f)
        wind_ms = round(vmax_f * 0.514444, 2) if vmax_f > 0 else None
        pt: dict = {
            "lat": round(lat_f, 4),
            "lng": round(lon_f, 4),
            "time": f"{int(year):04d}-{int(month):02d}-{int(day):02d}T{int(hour):02d}:00:00Z",
            "pressureHpa": to_float(pres),
        }
        if wind_ms is not None:
            pt["windMs"] = wind_ms
        path.append(pt)
    return path, peak_kt


def find_key(fields: list[str] | None, *needles: str) -> str | None:
    if not fields:
        return None
    for f in fields:
        for n in needles:
            if n.lower() in f.lower() or n in f:
                return f
    return None


def parse_half_deg(wkt: str | None) -> float:
    if not wkt:
        return GRID_HALF_DEG
    nums = [float(x) for x in WKT_NUM.findall(wkt)]
    # POLYGON: lon,lat pairs
    if len(nums) >= 8:
        lons = nums[0::2][:4]
        lats = nums[1::2][:4]
        if lons and lats:
            return max((max(lons) - min(lons)) / 2, (max(lats) - min(lats)) / 2, 1e-6)
    return GRID_HALF_DEG


def aggregate_exposure() -> tuple[list[dict], list[dict], dict, float]:
    """
    返回:
    - regions: 市町村聚合（JPY）
    - grid_cells: 有损失的网格点（JPY）
    - structure: 结构损失份额（百万 USD 演示口径，由 JPY 粗换算）
    - half_deg: 网格半宽
    """
    by_muni: dict[str, dict] = defaultdict(
        lambda: {
            "population": 0.0,
            "pgaPop": 0.0,
            "windPop": 0.0,
            "lossJpy": 0.0,
            "cells": 0,
            "latSum": 0.0,
            "lonSum": 0.0,
        }
    )
    grid_cells: list[dict] = []
    half_samples: list[float] = []

    f, reader, enc = open_csv(EXPOSURE_CSV)
    print(f"exposure encoding={enc}")
    try:
        fields = list(reader.fieldnames or [])
        pga_key = find_key(fields, "[gal]", "加速度") or fields[13]
        wind_key = find_key(fields, "风速", "風速", "m/s") or fields[14]
        loss_key = "Coupled_Loss_with_Other_JPY"
        if loss_key not in fields:
            raise KeyError(f"缺少列 {loss_key}: {fields}")

        for row in reader:
            muni = (row.get("ADM2_JA") or row.get("laa") or "未知").strip()
            pop = to_float(row.get("population")) or 0.0
            pga = to_float(row.get(pga_key)) or 0.0
            wind = to_float(row.get(wind_key)) or 0.0
            lat = to_float(row.get("latitude"))
            lon = to_float(row.get("longitude"))
            loss_jpy = to_float(row.get(loss_key)) or 0.0

            bucket = by_muni[muni]
            bucket["population"] += pop
            bucket["pgaPop"] += pga * max(pop, 1.0)
            bucket["windPop"] += wind * max(pop, 1.0)
            bucket["lossJpy"] += loss_jpy
            bucket["cells"] += 1
            if lat is not None and lon is not None:
                bucket["latSum"] += lat
                bucket["lonSum"] += lon

            if lat is None or lon is None:
                continue
            # 仅输出有损失的网格，避免 12 万空网格拖垮前端
            if loss_jpy <= 0:
                continue

            half = parse_half_deg(row.get("WKT"))
            half_samples.append(half)
            cell = {
                "lat": round(lat, 5),
                "lng": round(lon, 5),
                "lossJpy": round(loss_jpy, 2),
                "population": int(round(pop)),
                "name": muni,
            }
            pga_v = to_float(row.get(pga_key))
            wind_v = to_float(row.get(wind_key))
            if pga_v is not None:
                cell["pgaGal"] = round(pga_v, 3)
            if wind_v is not None:
                cell["windMs"] = round(wind_v, 3)
            grid_cells.append(cell)
    finally:
        f.close()

    half_deg = (
        sum(half_samples) / len(half_samples) if half_samples else GRID_HALF_DEG
    )

    regions = []
    for name, b in by_muni.items():
        weight = max(b["population"], b["cells"])
        mean_pga = b["pgaPop"] / weight if weight else 0.0
        mean_wind = b["windPop"] / weight if weight else 0.0
        regions.append(
            {
                "id": name,
                "name": name,
                "population": int(round(b["population"])),
                "meanPgaGal": round(mean_pga, 3),
                "meanWindMs": round(mean_wind, 3),
                "lossIndex": int(round(b["lossJpy"])),
                "lossJpy": round(b["lossJpy"], 2),
                "lat": round(b["latSum"] / b["cells"], 4) if b["cells"] else None,
                "lng": round(b["lonSum"] / b["cells"], 4) if b["cells"] else None,
            }
        )

    regions.sort(key=lambda r: r["lossJpy"], reverse=True)
    grid_cells.sort(key=lambda c: c["lossJpy"], reverse=True)

    total_jpy = sum(c["lossJpy"] for c in grid_cells)
    # 图表区仍用百万美元量级演示份额；按约 150 JPY/USD 粗换算
    total_million_usd = max(1, int(round(total_jpy / 150 / 1_000_000)))
    structure = {key: int(round(total_million_usd * share)) for key, share in STRUCTURE_SHARE.items()}
    drift = total_million_usd - sum(structure.values())
    structure["wood"] += drift

    return regions, grid_cells, structure, half_deg


def main() -> None:
    eq = load_earthquake()
    path, peak_kt = load_typhoon_track()
    regions, grid_cells, structure, half_deg = aggregate_exposure()

    peak_ms = round(peak_kt * 0.514444, 2) if peak_kt > 0 else None
    if peak_ms is None and regions:
        peak_ms = max(r["meanWindMs"] for r in regions)

    near_hokkaido = [p for p in path if 40 <= p["lat"] <= 46 and 138 <= p["lng"] <= 148]
    coupling_pt = near_hokkaido[len(near_hokkaido) // 2] if near_hokkaido else (path[-1] if path else None)
    total_jpy = sum(c["lossJpy"] for c in grid_cells)

    # 紧凑网格：数组行 [lat, lng, lossJpy, name, pop, pga, wind]
    compact_cells = []
    for c in grid_cells:
        compact_cells.append(
            [
                c["lat"],
                c["lng"],
                int(round(c["lossJpy"])),
                c.get("name") or "",
                int(c.get("population") or 0),
                None if c.get("pgaGal") is None else round(c["pgaGal"], 2),
                None if c.get("windMs") is None else round(c["windMs"], 2),
            ]
        )

    grid_payload = {
        "version": 1,
        "unit": "JPY",
        "halfDeg": round(half_deg, 8),
        "count": len(compact_cells),
        "cells": compact_cells,
    }
    OUT_GRID.parent.mkdir(parents=True, exist_ok=True)
    OUT_GRID.write_text(json.dumps(grid_payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    payload = {
        "version": 4,
        "event": {
            "id": "HIS-2018-IBURI-JEBI",
            "label": "2018 M6.7 Hokkaido Eastern Iburi Earthquake and Typhoon Jebi (No. 21)",
            "year": eq["year"],
            "magnitude": eq["magnitude"],
            "depthKm": eq["depthKm"],
            "eqTime": eq["time"],
            "epicenter": {"lat": eq["lat"], "lng": eq["lon"]},
            "typhoonCode": "No. 21",
            "typhoonName": "Jebi",
            "windMs": peak_ms,
            "windSpeed": round((peak_ms or 0) * 3.6, 1),
            # Jebi (No. 21) peaked over Japan ~2018-09-04; Iburi EQ on 2018-09-06 → TY then EQ
            "couplingType": "Typhoon followed by Earthquake",
            "typhoonPath": path,
            "typhoonAtCoupling": None
            if coupling_pt is None
            else {"lat": coupling_pt["lat"], "lng": coupling_pt["lng"], "windMs": coupling_pt.get("windMs")},
            "descriptions": [
                "Case study: 2018 M6.7 Hokkaido Eastern Iburi Earthquake and Typhoon Jebi (No. 21).",
                "Grid losses from Coupled_Loss_with_Other_JPY (JPY); map colored by grid centers.",
                f"Focal depth ≈ {eq['depthKm']:.0f} km; track points {len(path)}; loss grids {len(grid_cells)} / municipalities {len(regions)}.",
            ],
            "structureLoss": structure,
            "regions": regions,
            "gridHalfDeg": round(half_deg, 8),
            "gridCellsUrl": "/data/risk-grid-cells.json",
            "gridCells": [],
            "totalCoupledLossJpy": round(total_jpy, 2),
            "lossByRegion": {r["id"]: r["lossJpy"] for r in regions},
            "lossByPrefecture": {r["id"]: r["lossJpy"] for r in regions},
        },
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    size_mb = OUT.stat().st_size / (1024 * 1024)
    grid_mb = OUT_GRID.stat().st_size / (1024 * 1024)
    print(f"wrote {OUT} ({size_mb:.2f} MB)")
    print(f"wrote {OUT_GRID} ({grid_mb:.2f} MB, {len(compact_cells)} cells)")
    print(
        f"regions={len(regions)} gridCells={len(grid_cells)} "
        f"path={len(path)} peakWindMs={peak_ms} totalJpy={total_jpy:.0f}"
    )


if __name__ == "__main__":
    main()
